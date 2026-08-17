# -*- coding: utf-8 -*-
"""
总控脚本：运行 P1/P2/P3，输出 result1/2/3.xlsx、summary_all.json 与全部图
"""
import sys, os, json, math, time
sys.path.insert(0, r'D:/A题_work')
import numpy as np
import openpyxl
from solve3 import (load_case, dist_matrix, solve, route_total, mst_time, SERVICE)
from p3b import solve_p3, route_finish, START_H
from zones import load_zones

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import font_manager, patches

for f in ['Microsoft YaHei', 'SimHei', 'SimSun']:
    try:
        font_manager.findfont(f, fallback_to_default=False)
        plt.rcParams['font.sans-serif'] = [f]
        break
    except Exception:
        continue
plt.rcParams['axes.unicode_minus'] = False

RES = r'D:/A题_work/结果文件'
FIG = r'D:/A题_work/figures'
os.makedirs(RES, exist_ok=True)
os.makedirs(FIG, exist_ok=True)

CASES = ['Case1', 'Case2', 'Case3', 'Case4']
NMAP = {'Case1': 4, 'Case2': 2, 'Case3': 4, 'Case4': 4}

def save_result_xlsx(name, all_sols, all_times, path, p3=False):
    """all_sols: {case: [route,...]}; all_times: {case: [(arr,leave),...] per route}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for c in CASES:
        sol = all_sols[c]
        ws = wb.create_sheet(c)
        maxlen = max((len(r) for r in sol), default=0)
        ws.append(['UAV ID'] + [f'{k}th Inspection Point' for k in range(1, maxlen + 1)])
        for i, r in enumerate(sol):
            ws.append([i + 1] + [int(x) for x in r])
        ws2 = wb.create_sheet(c + '_Time')
        ws2.append(['UAV ID', 'Seq', 'Point_ID', 'Arrive_Time', 'Leave_Time'])
        for i, r in enumerate(sol):
            for k, (p, tarr, tleave) in enumerate(all_times[c][i]):
                ws2.append([i + 1, k + 1, int(p), round(tarr, 4), round(tleave, 4)])
    wb.save(path)

def times_from_sol(sol, D, start_h):
    out = {}
    for i, r in enumerate(sol):
        seq = []
        t = start_h
        prev = 0
        for p in r:
            t += D[prev, p]
            seq.append((p, t, t + SERVICE))
            t += SERVICE
            prev = p
        out[i] = seq
    return out

def plot_routes(case, sols, title, fname, zones=None, coords=None):
    fig, ax = plt.subplots(figsize=(7.5, 7.5))
    colors = plt.cm.tab10(np.linspace(0, 1, len(sols)))
    for i, r in enumerate(sols):
        xy = np.array([[0, 0]] + [coords[p] for p in r] + [[0, 0]])
        ax.plot(xy[:, 0], xy[:, 1], '-o', color=colors[i], lw=1.2, ms=2.5,
                label=f'UAV{i+1} ({len(r)}点)')
    if zones:
        for z in zones:
            c = patches.Circle(z['c'], z['r'], fill=False, edgecolor='red',
                               lw=1.2, ls='--', alpha=0.8)
            ax.add_patch(c)
            ax.text(z['c'][0], z['c'][1], z['id'], color='red', fontsize=8,
                    ha='center', va='center')
    ax.scatter(0, 0, marker='*', c='k', s=180, zorder=6, label='基地')
    ax.set_title(title)
    ax.set_xlabel('X (单位=100m)')
    ax.set_ylabel('Y (单位=100m)')
    ax.legend(fontsize=7, loc='best')
    ax.set_aspect('equal')
    plt.tight_layout()
    plt.savefig(fname, dpi=130)
    plt.close()

def main():
    summary = {}
    # ---------- P1 & P2 ----------
    for c in CASES:
        pts = load_case(c)
        D = dist_matrix(pts)
        M = sum(p[3] for p in pts)
        S = M * SERVICE
        T_mst = mst_time(D)
        Nmin = NMAP[c]
        sol1, obj1, D = solve(pts, Nmin, n_restart=8, max_iter=1500, seed=21, objective='makespan')
        sol2, obj2, D = solve(pts, Nmin, n_restart=6, max_iter=1500, seed=21,
                              init_sol=sol1, objective='balance')
        if obj2[0] > obj1[0] + 1e-6:
            sol2, obj2 = sol1, obj1
        l1 = [route_total(r, D) for r in sol1]
        l2 = [route_total(r, D) for r in sol2]
        summary[c] = dict(n_points=len(pts), visits=M, S_h=round(S, 4),
                          MST_h=round(T_mst, 4), Nmin=Nmin,
                          lb1=math.ceil(S / 9), lb2=math.ceil((S + T_mst) / 9),
                          P1=dict(N=Nmin, Tmax=round(float(obj1[0]), 4), Tmin=round(float(min(l1)), 4),
                                  loads=[round(float(x), 4) for x in l1]),
                          P2=dict(N=Nmin, Tmax=round(float(obj2[0]), 4), Tmin=round(float(min(l2)), 4),
                                  delta=round(float(obj2[1]), 4), loads=[round(float(x), 4) for x in l2]))
        summary[c]['P1_sol'] = [list(map(int, r)) for r in sol1]
        summary[c]['P2_sol'] = [list(map(int, r)) for r in sol2]
        print(f'P1/P2 {c} done: Nmin={Nmin} P1={obj1[0]:.3f} P2(delta)={obj2[1]:.3f}', flush=True)

    # ---------- P3 ----------
    for c in CASES:
        pts = load_case(c)
        zones = load_zones(c)
        D = dist_matrix(pts)
        orders, rts, wait, tl, D0, zs, coords, pts2 = solve_p3(c, NMAP[c])
        tmax = max(rts); tmin = min(rts)
        summary[c]['P3'] = dict(N=NMAP[c], Tmax=round(float(tmax), 4), Tmin=round(float(tmin), 4),
                                delta=round(float(tmax - tmin), 4),
                                route_times=[round(float(x), 4) for x in rts],
                                wait_h=round(float(wait), 4))
        summary[c]['P3_sol'] = [list(map(int, o)) for o in orders]
        summary[c]['P3_times'] = tl
        print(f'P3 {c} done: Tmax={tmax:.3f} delta={tmax-tmin:.3f} wait={wait:.3f}', flush=True)

    # ---------- 写 result xlsx ----------
    # P1
    s1 = {c: summary[c]['P1_sol'] for c in CASES}
    t1 = {c: times_from_sol(summary[c]['P1_sol'], dist_matrix(load_case(c)), 0.0) for c in CASES}
    save_result_xlsx('result1', s1, t1, os.path.join(RES, 'result1.xlsx'))
    # P2
    s2 = {c: summary[c]['P2_sol'] for c in CASES}
    t2 = {c: times_from_sol(summary[c]['P2_sol'], dist_matrix(load_case(c)), 0.0) for c in CASES}
    save_result_xlsx('result2', s2, t2, os.path.join(RES, 'result2.xlsx'))
    # P3 (含时间)
    s3 = {c: summary[c]['P3_sol'] for c in CASES}
    t3 = {}
    for c in CASES:
        t3[c] = {i: [(p, t, t + SERVICE) for (p, t) in summary[c]['P3_times'][i]]
                 for i in range(len(summary[c]['P3_times']))}
    save_result_xlsx('result3', s3, t3, os.path.join(RES, 'result3.xlsx'), p3=True)

    # ---------- 图 ----------
    for c in CASES:
        pts = load_case(c)
        coords = [None] + [(p[1], p[2]) for p in pts]
        zones = load_zones(c)
        # P1 路线
        plot_routes(c, summary[c]['P1_sol'],
                    f'{c} 问题一 路径规划 (N={summary[c]["Nmin"]}, Tmax={summary[c]["P1"]["Tmax"]}h)',
                    os.path.join(FIG, f'{c}_P1.png'), coords=coords)
        # P3 路线 + 禁飞区
        plot_routes(c, summary[c]['P3_sol'],
                    f'{c} 问题三 路径规划(含禁飞区) (N={summary[c]["P3"]["N"]}, Tmax={summary[c]["P3"]["Tmax"]}h)',
                    os.path.join(FIG, f'{c}_P3.png'), zones=zones, coords=coords)
        # 负载条形图
        fig, ax = plt.subplots(figsize=(7, 3.2))
        x = np.arange(summary[c]['Nmin'])
        w = 0.3
        ax.bar(x - w, summary[c]['P1']['loads'], w, label='问题一')
        ax.bar(x, summary[c]['P2']['loads'], w, label='问题二')
        ax.bar(x + w, summary[c]['P3']['route_times'], w, label='问题三')
        ax.axhline(9, color='r', ls='--', lw=1, label='9小时')
        ax.set_xticks(x)
        ax.set_xticklabels([f'UAV{i+1}' for i in x])
        ax.set_ylabel('工作时间 (h)')
        ax.set_title(f'{c} 无人机负载对比')
        ax.legend(fontsize=8)
        plt.tight_layout()
        plt.savefig(os.path.join(FIG, f'{c}_loads.png'), dpi=130)
        plt.close()
        print(f'figs {c} done', flush=True)

    # 保存 summary（去掉sol等大对象）
    clean = {}
    for c in CASES:
        clean[c] = {k: v for k, v in summary[c].items() if not k.endswith('_sol') and k != 'P3_times'}
    json.dump(clean, open(r'D:/A题_work/summary_all.json', 'w', encoding='utf-8'),
              ensure_ascii=False, indent=1)
    print('ALL DONE')
    print(json.dumps(clean, ensure_ascii=False, indent=1))

if __name__ == '__main__':
    main()
