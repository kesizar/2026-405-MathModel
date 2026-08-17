# -*- coding: utf-8 -*-
"""
A题 多无人机协同巡检路径优化 —— 主求解器
数据加载 + 问题1/2/3 求解 + 结果输出 + 绘图
"""
import numpy as np
import openpyxl
import json
import os
import math
from collections import Counter

BASE = r'D:/HuaweiMoveData/Users/17669/Desktop/A 题'
OUT = r'D:/A题_work'

SCALE = 0.1      # km per unit
SPEED = 55.0     # km/h
SERVICE = 5.0 / 60.0   # hours per visit

LVL = {'I': 3, 'II': 2, 'III': 1}

# ---------------- load data ----------------
def load_case(sheet):
    wb = openpyxl.load_workbook(os.path.join(BASE, '附件1.xlsx'), data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    pts = []  # (id, x, y, demand)
    for r in rows[1:]:
        if r[0] is None:
            continue
        pid, x, y, lvl = int(r[0]), float(r[1]), float(r[2]), str(r[3]).strip()
        pts.append((pid, x, y, LVL[lvl]))
    return pts

def load_zones(sheet):
    wb = openpyxl.load_workbook(os.path.join(BASE, '附件2.xlsx'), data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    zones = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        zid = str(r[0]).strip()
        cx, cy, rad = float(r[1]), float(r[2]), float(r[3])
        st, en = str(r[4]), str(r[5])
        zones.append((zid, cx, cy, rad, st, en))
    return zones

def time_to_h(tstr):
    h, m = tstr.strip().split(':')
    return int(h) + int(m) / 60.0

# ---------------- distance / time ----------------
def dist_matrix(pts):
    """pts: list of (id,x,y,demand); returns travel-time matrix in hours between point coords (depot idx 0)."""
    coords = np.array([[0.0, 0.0]] + [[p[1], p[2]] for p in pts])
    d = np.sqrt(((coords[:, None, :] - coords[None, :, :]) ** 2).sum(-1))
    return d * SCALE / SPEED  # hours

# ---------------- statistics ----------------
def stats(pts):
    M = sum(p[3] for p in pts)
    S_h = M * SERVICE
    return len(pts), M, S_h

if __name__ == '__main__':
    for c in ['Case1', 'Case2', 'Case3', 'Case4']:
        pts = load_case(c)
        n, M, S = stats(pts)
        lvl = Counter(p[3] for p in pts)
        print(f'{c}: points={n}, visits={M}, inspection_h={S:.3f}, '
              f'level_counts(I,II,III)={lvl.get(3,0)},{lvl.get(2,0)},{lvl.get(1,0)}')
        print(f'  ceil(S/9h) = {math.ceil(S/9.0)}')
        z = load_zones(c)
        print(f'  zones={len(z)}')
